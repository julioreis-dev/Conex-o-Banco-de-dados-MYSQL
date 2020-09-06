from openpyxl import load_workbook
import mysql.connector
import test_connection


def conectar_bd(*args):
    dict_gerais = args[1]
    connection = mysql.connector.connect(host=args[2], database=args[3], user=args[4], password=args[5], port=args[6])
    for contrato_icj in args[0]:
        extracao = dict_gerais[contrato_icj]
        sql = 'INSERT INTO contratos (icj, equipamento, embarcacao, empresa, inicio, termino, cessão) ' \
              'VALUES (%s, %s, %s, %s, %s, %s, %s)'
        values = (contrato_icj, extracao[0], extracao[1], extracao[2], extracao[3], extracao[4], extracao[5])
        cursor = connection.cursor()
        cursor.execute(sql, values)
        #connection.commit()
        print(cursor.rowcount, f'Registro do ICJ {contrato_icj} inserido com sucesso.')
    connection.close()


def catalogar(*args):
    contador = args[0].max_row
    dict_dados = {}
    lista_dados = []
    lista_icj = []
    for linha in range(2, contador + 1):
        icj = args[0].cell(row=linha, column=1).value
        lista_icj.append(icj)
        equipamento = args[0].cell(row=linha, column=2).value
        lista_dados.append(equipamento)
        embarcacao = args[0].cell(row=linha, column=4).value
        lista_dados.append(embarcacao)
        empresa = args[0].cell(row=linha, column=9).value
        lista_dados.append(empresa)
        inicio = args[0].cell(row=linha, column=10).value
        lista_dados.append(inicio)
        termino = args[0].cell(row=linha, column=11).value
        lista_dados.append(termino)
        cessao = args[0].cell(row=linha, column=12).value
        lista_dados.append(cessao)
        dict_dados[icj] = lista_dados
        dict_dados.copy()
        lista_dados = []
    return lista_icj, dict_dados


def main():
    wb = load_workbook(filename='Planilha Guia_dados.xlsx')
    sheet = wb['Info Contrato']
    teste = test_connection.testar_bd('localhost', 'frota_maritima', 'root', '', 3306)
    if teste == 'ok':
        dados_contratual = catalogar(sheet)
        print('Teste de conexão com o banco de dados estabelecida com sucesso!')
        nova_lista = sorted(set(dados_contratual[0]))
        conectar_bd(nova_lista, dados_contratual[1], 'localhost', 'frota_maritima', 'root', '', 3306)
    else:
        print('Conexão não estabelecida!')
        exit()


if __name__ == '__main__':
    main()
